import openpyxl as xl

# # Load workbook
wb = xl.load_workbook("transactions.xlsx")

# # Select active sheet
sheet = wb['Sheet1']
cell=sheet['a1']
cell=sheet.cell(1,1)
print(cell.value)
for row in range(2, sheet.max_row+1):
    cell=sheet.cell(row,3)
    print(cell.value)
    disscount=cell.value-cell.value*0.10
    disscount_price=sheet.cell(row,4)
    disscount_price.value=disscount
    wb.save('transactions2.xlsx')
     
